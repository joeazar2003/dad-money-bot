import os
import re
import io
import hmac
from collections import defaultdict
from datetime import datetime

import requests
from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

app = Flask(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
# Your personal Telegram numeric chat ID. Only messages from this chat are logged.
ALLOWED_CHAT_ID = int(os.environ["ALLOWED_CHAT_ID"])

API_URL = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
XLSX_PATH = "dad_money.xlsx"
HOME_CURRENCY = os.environ.get("HOME_CURRENCY", "CAD").upper()
PAP_API_KEY = os.environ.get("PAP_API_KEY")

# --- Google Drive backup ---
# Render's free plan wipes local disk on every restart/redeploy, so the
# Excel file is mirrored to Google Drive after every change and restored
# from there on startup if the local copy is missing.
GOOGLE_CREDS_PATH = os.environ.get("GOOGLE_CREDS_PATH", "/etc/secrets/google-credentials.json")
DRIVE_FILE_NAME = "dad_money.xlsx"
DRIVE_SHARE_EMAIL = os.environ.get("DRIVE_SHARE_EMAIL", "joeazar2003@gmail.com")
_drive_service = None
_drive_file_id = None

# --- Finance tracker Google Sheet ---
# Step-by-step flow (/log) that appends a row to the "Joe Finance Tracker"
# Sheet, pulling the PAP (dad's money) figure straight from get_total()
# above. Layout: Safe and Drawer are always asked; the people who owe you
# or are owed money change every time, so those are collected as one
# free-form "anyone else" answer (amount + name pairs) rather than fixed
# columns; then the three bank balances, in order.
FINANCE_SPREADSHEET_ID = os.environ.get(
    "FINANCE_SPREADSHEET_ID", "15RlB-83RQvTVECBmmcXsPYDTirGhIbzmwe3yVkYUHlA"
)
FINANCE_SHEET_NAME = os.environ.get("FINANCE_SHEET_NAME", "Sheet1")
# The new per-entry "block" layout (auto date, TD Credit/Plat/Aero, colored
# groups) is written to this tab. Kept separate from FINANCE_SHEET_NAME so
# the old row-based Sheet1 data isn't touched until it's ready to migrate.
FINANCE_BLOCK_SHEET_NAME = os.environ.get("FINANCE_BLOCK_SHEET_NAME", "Layout Draft")
FINANCE_BANK_ORDER = ["TD", "RBC", "Scotiabank"]
FINANCE_BANK_LINKS = {
    "TD": "https://easyweb.td.com/waw/idp/login.htm",
    "RBC": "https://www1.royalbank.com/sgw3/secureauth/login",
    "Scotiabank": "https://www1.scotiaonline.scotiabank.com/online/authentication/authentication.bns",
}
# (session key, display label) for the credit-card question sequence.
FINANCE_CREDIT_ORDER = [("td_credit", "TD Credit"), ("plat", "Plat"), ("aero", "Aero")]
_sheets_service = None
finance_sessions = {}


def get_drive_service():
    global _drive_service
    if _drive_service is None:
        creds = service_account.Credentials.from_service_account_file(
            GOOGLE_CREDS_PATH,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        _drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
    return _drive_service


def find_drive_file_id():
    global _drive_file_id
    if _drive_file_id:
        return _drive_file_id
    try:
        service = get_drive_service()
        resp = service.files().list(
            q=f"name='{DRIVE_FILE_NAME}' and trashed=false",
            spaces="drive",
            fields="files(id, name)",
        ).execute()
        files = resp.get("files", [])
        if files:
            _drive_file_id = files[0]["id"]
    except Exception as e:
        print("Drive lookup failed:", e)
    return _drive_file_id


def download_from_drive():
    try:
        file_id = find_drive_file_id()
        if not file_id:
            return False
        service = get_drive_service()
        req = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        with open(XLSX_PATH, "wb") as f:
            f.write(buf.getvalue())
        return True
    except Exception as e:
        print("Drive download failed:", e)
        return False


def upload_to_drive():
    global _drive_file_id
    try:
        service = get_drive_service()
        file_id = find_drive_file_id()
        media = MediaFileUpload(
            XLSX_PATH,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False,
        )
        if file_id:
            service.files().update(fileId=file_id, media_body=media).execute()
        else:
            created = service.files().create(
                body={"name": DRIVE_FILE_NAME}, media_body=media, fields="id"
            ).execute()
            _drive_file_id = created["id"]
            if DRIVE_SHARE_EMAIL:
                try:
                    service.permissions().create(
                        fileId=_drive_file_id,
                        body={"type": "user", "role": "writer", "emailAddress": DRIVE_SHARE_EMAIL},
                        sendNotificationEmail=False,
                    ).execute()
                except Exception as e:
                    print("Drive share failed:", e)
    except Exception as e:
        print("Drive upload failed:", e)


def get_sheets_service():
    global _sheets_service
    if _sheets_service is None:
        creds = service_account.Credentials.from_service_account_file(
            GOOGLE_CREDS_PATH,
            scopes=["https://www.googleapis.com/auth/spreadsheets"],
        )
        _sheets_service = build("sheets", "v4", credentials=creds, cache_discovery=False)
    return _sheets_service


def parse_finance_amount(text):
    """Returns a float, or None if it couldn't be parsed. 'skip'/'0'/etc mean 0."""
    low = text.strip().lower()
    if low in ("skip", "0", "-", "none", "n/a", ""):
        return 0.0
    cleaned = text.strip().replace(",", "").replace("$", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


NUMBER_TOKEN_RE = re.compile(r"^[+-]?\d+(?:\.\d+)?$")


def parse_others(text):
    """Parses a free-form list like '299 george -100 bassel - 144 jose 350 jaber'
    into [(amount, name), ...]. Positive = owed to you, negative = you owe them."""
    normalized = re.sub(r"-\s+(\d)", r"-\1", text.replace(",", " "))
    tokens = normalized.split()
    items = []
    i, n = 0, len(tokens)
    while i < n:
        tok = tokens[i]
        if NUMBER_TOKEN_RE.match(tok):
            amount = float(tok)
            i += 1
            name_parts = []
            while i < n and not NUMBER_TOKEN_RE.match(tokens[i]):
                name_parts.append(tokens[i])
                i += 1
            name = " ".join(name_parts).strip(" ,") or "Unknown"
            items.append((amount, name))
        else:
            i += 1
    return items


def format_others_detail(items):
    if not items:
        return "(none)"
    parts = []
    for amt, name in items:
        sign = "+" if amt >= 0 else "-"
        val = abs(amt)
        val_str = f"{val:.0f}" if val == int(val) else f"{val:.2f}"
        parts.append(f"{sign}{val_str} {name}")
    return ", ".join(parts)


def bank_keyboard(bank):
    url = FINANCE_BANK_LINKS.get(bank)
    if not url:
        return None
    return {"inline_keyboard": [[{"text": f"Open {bank}", "url": url}]]}


def sheet_link_keyboard(url):
    return {"inline_keyboard": [[{"text": "\U0001F4C4 Open in Sheet", "url": url}]]}


_sheet_id_cache = {}


def get_sheet_id(sheet_name):
    """Look up the numeric sheetId (gid) for a tab by its title."""
    if sheet_name in _sheet_id_cache:
        return _sheet_id_cache[sheet_name]
    service = get_sheets_service()
    meta = service.spreadsheets().get(spreadsheetId=FINANCE_SPREADSHEET_ID).execute()
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if props.get("title") == sheet_name:
            _sheet_id_cache[sheet_name] = props.get("sheetId")
            return props.get("sheetId")
    raise ValueError(f"Sheet tab '{sheet_name}' not found")


def find_block_start_row(sheet_name):
    """First row to start a new day's block: one blank spacer row below
    whatever is already there, or row 1 if the tab is empty."""
    service = get_sheets_service()
    resp = service.spreadsheets().values().get(
        spreadsheetId=FINANCE_SPREADSHEET_ID,
        range=f"{sheet_name}!A:K",
    ).execute()
    rows = resp.get("values", [])
    last_used = 0
    for i, row in enumerate(rows, start=1):
        if any(str(v).strip() for v in row):
            last_used = i
    return 1 if last_used == 0 else last_used + 2


def _rgb(r, g, b):
    return {"red": r, "green": g, "blue": b}


WHITE = _rgb(1, 1, 1)
PASTEL_ITEM_COLORS = [
    _rgb(0.80, 0.86, 0.97),  # blue   (Safe)
    _rgb(0.82, 0.93, 0.82),  # green  (Drawer)
    _rgb(0.99, 0.94, 0.78),  # yellow
    _rgb(0.87, 0.82, 0.94),  # purple
    _rgb(0.99, 0.85, 0.85),  # pink
    _rgb(0.80, 0.93, 0.93),  # teal
]
BANK_COLUMN_COLORS = {
    "TD": _rgb(0.82, 0.93, 0.82),          # green
    "RBC": _rgb(0.80, 0.86, 0.97),         # blue
    "Scotiabank": _rgb(0.98, 0.82, 0.82),  # red/pink
}
CREDIT_ROW_COLORS = {
    "TD Credit": _rgb(0.82, 0.93, 0.82),  # green
    "Plat": _rgb(0.99, 0.94, 0.78),       # yellow
    "Aero": _rgb(0.80, 0.86, 0.97),       # blue
}
TOTALS_COLOR = _rgb(0.85, 0.85, 0.85)  # light gray

THIN_BLACK_BORDER = {"style": "SOLID", "width": 1, "color": {"red": 0, "green": 0, "blue": 0}}
THICK_BLACK_BORDER = {"style": "SOLID_THICK", "width": 2, "color": {"red": 0, "green": 0, "blue": 0}}


def _cell(value=None, bold=False, bg=None, align="CENTER", font_size=10):
    fmt = {
        "textFormat": {"bold": bold, "fontSize": font_size},
        "horizontalAlignment": align,
        "verticalAlignment": "MIDDLE",
    }
    if bg:
        fmt["backgroundColor"] = bg
    cell = {"userEnteredFormat": fmt}
    if value is not None:
        if isinstance(value, (int, float)):
            cell["userEnteredValue"] = {"numberValue": value}
        else:
            cell["userEnteredValue"] = {"stringValue": str(value)}
    return cell


def _formula_cell(formula, bold=False, bg=None, align="CENTER", font_size=10):
    fmt = {
        "textFormat": {"bold": bold, "fontSize": font_size},
        "horizontalAlignment": align,
        "verticalAlignment": "MIDDLE",
    }
    if bg:
        fmt["backgroundColor"] = bg
    return {"userEnteredFormat": fmt, "userEnteredValue": {"formulaValue": formula}}


def build_finance_block_rows(session, start_row):
    """Build the 11-column (A:K) grid of cells for one day's block, matching
    the design confirmed live in the Layout Draft tab. Subtotal and Net are
    written as live Sheets formulas (not baked-in numbers) so that editing
    any source cell directly in the Sheet - Safe, a bank balance, a credit
    card, an "others" amount - recalculates them automatically."""
    date_label = session["date_display"]
    v = session["values"]
    others = session["others_items"]  # list of (amount, name)
    pap = session["pap"]

    n_rows = max(3, 2 + len(others))
    r0, r1, r2 = start_row, start_row + 1, start_row + 2
    end_row = start_row + n_rows - 1
    subtotal_formula = f"=C{r0}+C{r1}+SUM(C{r2}:C{end_row})+D{r1}+E{r1}+F{r1}+G{r0}+G{r1}+G{r2}"
    net_formula = f"=I{r1}-J{r1}"
    grid = [[None] * 11 for _ in range(n_rows)]  # columns A..K -> index 0..10

    color_i = 0

    def next_item_color():
        nonlocal color_i
        c = PASTEL_ITEM_COLORS[color_i % len(PASTEL_ITEM_COLORS)]
        color_i += 1
        return c

    safe_color = next_item_color()
    drawer_color = next_item_color()

    # Row 0: date (title), Safe, TD Credit
    grid[0][0] = _cell(date_label, bold=True, align="LEFT", font_size=14, bg=WHITE)
    grid[0][1] = _cell("Safe", bold=True, align="RIGHT", bg=safe_color)
    grid[0][2] = _cell(v["Safe"], bg=safe_color)
    grid[0][6] = _cell(v["td_credit"], bg=CREDIT_ROW_COLORS["TD Credit"])
    grid[0][7] = _cell("TD Credit", bold=True, align="LEFT", bg=CREDIT_ROW_COLORS["TD Credit"])

    # Row 1: Drawer, bank values, Plat, totals values
    grid[1][1] = _cell("Drawer", bold=True, align="RIGHT", bg=drawer_color)
    grid[1][2] = _cell(v["Drawer"], bg=drawer_color)
    grid[1][3] = _cell(v["TD"], bg=BANK_COLUMN_COLORS["TD"])
    grid[1][4] = _cell(v["RBC"], bg=BANK_COLUMN_COLORS["RBC"])
    grid[1][5] = _cell(v["Scotiabank"], bg=BANK_COLUMN_COLORS["Scotiabank"])
    grid[1][6] = _cell(v["plat"], bg=CREDIT_ROW_COLORS["Plat"])
    grid[1][7] = _cell("Plat", bold=True, align="LEFT", bg=CREDIT_ROW_COLORS["Plat"])
    grid[1][8] = _formula_cell(subtotal_formula, bg=TOTALS_COLOR)
    grid[1][9] = _cell(pap, bg=TOTALS_COLOR)
    grid[1][10] = _formula_cell(net_formula, bg=TOTALS_COLOR)

    # Row 2: first "other" (if any), bank labels, Aero, totals labels
    grid[2][3] = _cell("TD", bold=True, bg=BANK_COLUMN_COLORS["TD"])
    grid[2][4] = _cell("RBC", bold=True, bg=BANK_COLUMN_COLORS["RBC"])
    grid[2][5] = _cell("SCO", bold=True, bg=BANK_COLUMN_COLORS["Scotiabank"])
    grid[2][6] = _cell(v["aero"], bg=CREDIT_ROW_COLORS["Aero"])
    grid[2][7] = _cell("Aero", bold=True, align="LEFT", bg=CREDIT_ROW_COLORS["Aero"])
    grid[2][8] = _cell("Subtotal", bold=True, bg=TOTALS_COLOR)
    grid[2][9] = _cell("PAP", bold=True, bg=TOTALS_COLOR)
    grid[2][10] = _cell("Net", bold=True, bg=TOTALS_COLOR)

    for i, (amount, name) in enumerate(others):
        row_idx = 2 + i
        color = next_item_color()
        grid[row_idx][1] = _cell(name, bold=True, align="RIGHT", bg=color)
        grid[row_idx][2] = _cell(amount, bg=color)

    return grid, n_rows


def write_finance_block(session):
    """Write one full day's block to FINANCE_BLOCK_SHEET_NAME, stacked below
    whatever is already there (with a blank spacer row), fully formatted
    (colors, bold labels, alignment, grid + thick outer border) in one
    batchUpdate call. Returns the 1-based row the block starts on."""
    sheet_name = FINANCE_BLOCK_SHEET_NAME
    sheet_id = get_sheet_id(sheet_name)
    start_row = find_block_start_row(sheet_name)  # 1-based
    grid, n_rows = build_finance_block_rows(session, start_row)

    start_row_index = start_row - 1
    end_row_index = start_row_index + n_rows

    rows_payload = []
    for r in range(n_rows):
        row_values = []
        for c in range(11):
            cell = grid[r][c]
            row_values.append(cell if cell is not None else {"userEnteredFormat": {"verticalAlignment": "MIDDLE"}})
        rows_payload.append({"values": row_values})

    block_range = {
        "sheetId": sheet_id,
        "startRowIndex": start_row_index,
        "endRowIndex": end_row_index,
        "startColumnIndex": 0,
        "endColumnIndex": 11,
    }
    requests_body = [
        {
            "updateCells": {
                "rows": rows_payload,
                "fields": "userEnteredValue,userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)",
                "range": block_range,
            }
        },
        {
            "updateBorders": {
                "range": block_range,
                "top": THIN_BLACK_BORDER, "bottom": THIN_BLACK_BORDER,
                "left": THIN_BLACK_BORDER, "right": THIN_BLACK_BORDER,
                "innerHorizontal": THIN_BLACK_BORDER, "innerVertical": THIN_BLACK_BORDER,
            }
        },
        {
            "updateBorders": {
                "range": block_range,
                "top": THICK_BLACK_BORDER, "bottom": THICK_BLACK_BORDER,
                "left": THICK_BLACK_BORDER, "right": THICK_BLACK_BORDER,
            }
        },
    ]

    service = get_sheets_service()
    service.spreadsheets().batchUpdate(
        spreadsheetId=FINANCE_SPREADSHEET_ID, body={"requests": requests_body}
    ).execute()
    return start_row, sheet_id, n_rows


def sheet_block_url(sheet_id, start_row, n_rows):
    """A link that opens the Sheet (web or the mobile app) with the just-written
    block's range selected/highlighted, so a phone user can actually see it
    instead of just getting a text confirmation."""
    end_row = start_row + n_rows - 1
    return (
        f"https://docs.google.com/spreadsheets/d/{FINANCE_SPREADSHEET_ID}"
        f"/edit#gid={sheet_id}&range=A{start_row}:K{end_row}"
    )


def write_single_cell(sheet_name, cell_a1, value):
    """Write one value directly into one cell, e.g. to correct a number after
    the fact. USER_ENTERED so a plain number lands as a number (and would
    still work if a formula string were ever passed in)."""
    service = get_sheets_service()
    service.spreadsheets().values().update(
        spreadsheetId=FINANCE_SPREADSHEET_ID,
        range=f"{sheet_name}!{cell_a1}",
        valueInputOption="USER_ENTERED",
        body={"values": [[value]]},
    ).execute()


COL_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5, "G": 6, "H": 7, "I": 8, "J": 9, "K": 10}

# (label, column, which block row it's on, is_credit_card)
# "is_credit_card" fields are stored as negative numbers, same as during /log.
EDIT_DIRECT_FIELDS = [
    ("Safe", "C", "r0", False),
    ("Drawer", "C", "r1", False),
    ("TD", "D", "r1", False),
    ("RBC", "E", "r1", False),
    ("Scotiabank", "F", "r1", False),
    ("TD Credit", "G", "r0", True),
    ("Plat", "G", "r1", True),
    ("Aero", "G", "r2", True),
]


def edit_field_cell(row_key, col, start_row):
    row = {"r0": start_row, "r1": start_row + 1, "r2": start_row + 2}[row_key]
    return f"{col}{row}"


def find_last_block(sheet_name):
    """Locate the most recently written block on the tab: its start row,
    row count, sheetId, date label, and a live A:K snapshot (read fresh, so
    any manual edits already made are reflected). Returns None if the tab
    has no blocks yet."""
    service = get_sheets_service()
    sheet_id = get_sheet_id(sheet_name)
    resp = service.spreadsheets().values().get(
        spreadsheetId=FINANCE_SPREADSHEET_ID,
        range=f"{sheet_name}!A:K",
    ).execute()
    rows = resp.get("values", [])

    block_starts = [i + 1 for i, row in enumerate(rows) if row and str(row[0]).strip()]
    if not block_starts:
        return None
    start_row = block_starts[-1]

    n_rows = 0
    for row in rows[start_row - 1:]:
        if not any(str(v).strip() for v in row):
            break
        n_rows += 1

    end_row = start_row + n_rows - 1
    padded = []
    for row in rows[start_row - 1:end_row]:
        padded.append((list(row) + [""] * 11)[:11])
    while len(padded) < n_rows:
        padded.append([""] * 11)

    return {
        "start_row": start_row,
        "n_rows": n_rows,
        "sheet_id": sheet_id,
        "date_label": padded[0][0] or "(unknown date)",
        "values": padded,
    }


def block_others(block):
    """[(name, amount, row_index_within_block), ...] for the "others" rows
    (block row index 2 onward)."""
    items = []
    for row_idx in range(2, block["n_rows"]):
        name = block["values"][row_idx][COL_INDEX["B"]]
        if str(name).strip():
            try:
                amt = float(block["values"][row_idx][COL_INDEX["C"]])
            except (TypeError, ValueError):
                amt = 0.0
            items.append((str(name).strip(), amt, row_idx))
    return items


def field_current_value(block, label, col, row_key):
    row_idx = {"r0": 0, "r1": 1, "r2": 2}[row_key]
    raw = block["values"][row_idx][COL_INDEX[col]]
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def edit_fields_prompt(block):
    lines = [f"Editing the {block['date_label']} entry (row {block['start_row']}). Reply with a number:"]
    for i, (label, col, row_key, _) in enumerate(EDIT_DIRECT_FIELDS, start=1):
        current = field_current_value(block, label, col, row_key)
        lines.append(f"{i}. {label} (${current:,.2f})")
    others = block_others(block)
    if others:
        lines.append(f"{len(EDIT_DIRECT_FIELDS) + 1}. Someone owes/is owed money ({', '.join(n for n, _, _ in others)})")
    lines.append("\n/cancel to stop.")
    return "\n".join(lines)


edit_sessions = {}


def start_edit_session(chat_id):
    block = find_last_block(FINANCE_BLOCK_SHEET_NAME)
    if not block:
        send_message(chat_id, "No entries in the sheet yet to edit.")
        return
    edit_sessions[chat_id] = {"step": "pick_field", "block": block}
    send_message(chat_id, edit_fields_prompt(block))


def _finish_edit(chat_id, session, label, addr, value_to_write):
    block = session["block"]
    try:
        write_single_cell(FINANCE_BLOCK_SHEET_NAME, addr, value_to_write)
    except Exception as e:
        print("Edit write failed:", e)
        send_message(chat_id, f"Couldn't write that to the Sheet ({e}). Try /edit again in a bit.")
        del edit_sessions[chat_id]
        return
    del edit_sessions[chat_id]
    url = sheet_block_url(block["sheet_id"], block["start_row"], block["n_rows"])
    send_message(
        chat_id,
        f"Updated {label} to ${value_to_write:,.2f} (cell {addr}). Subtotal and Net recalculate automatically.",
        reply_markup=sheet_link_keyboard(url),
    )


def handle_edit_session(chat_id, text):
    session = edit_sessions[chat_id]
    stripped = text.strip()

    if stripped.lower() == "/cancel":
        del edit_sessions[chat_id]
        send_message(chat_id, "Cancelled, nothing changed.")
        return

    block = session["block"]
    step = session["step"]

    if step == "pick_field":
        others = block_others(block)
        choice = stripped.lower()
        n_direct = len(EDIT_DIRECT_FIELDS)

        if choice.isdigit() and 1 <= int(choice) <= n_direct:
            field = EDIT_DIRECT_FIELDS[int(choice) - 1]
        else:
            field = next((f for f in EDIT_DIRECT_FIELDS if choice == f[0].lower()), None)

        if field is not None:
            label, col, row_key, is_credit = field
            addr = edit_field_cell(row_key, col, block["start_row"])
            current = field_current_value(block, label, col, row_key)
            session["step"] = "new_value"
            session["target"] = {"label": label, "addr": addr, "is_credit": is_credit}
            send_message(chat_id, f"New value for {label}? (currently ${current:,.2f}, cell {addr})")
            return

        if (choice.isdigit() and int(choice) == n_direct + 1 and others) or (
            others and any(choice == n.lower() for n, _, _ in others)
        ):
            if choice.isdigit():
                pick_lines = ["Which one? Reply with a number:"]
                for i, (name, amt, _) in enumerate(others, start=1):
                    pick_lines.append(f"{i}. {name} (${amt:,.2f})")
                session["step"] = "pick_other"
                session["others"] = others
                send_message(chat_id, "\n".join(pick_lines))
                return
            name, amt, row_idx = next(o for o in others if o[0].lower() == choice)
            addr = f"C{block['start_row'] + row_idx}"
            session["step"] = "new_value"
            session["target"] = {"label": name, "addr": addr, "is_credit": False}
            send_message(chat_id, f"New amount for {name}? (currently ${amt:,.2f}, cell {addr})")
            return

        send_message(chat_id, "Didn't catch that. Reply with one of the numbers above, or /cancel.")
        return

    if step == "pick_other":
        others = session["others"]
        match = None
        if stripped.isdigit() and 1 <= int(stripped) <= len(others):
            match = others[int(stripped) - 1]
        else:
            match = next((o for o in others if o[0].lower() == stripped.lower()), None)
        if match is None:
            send_message(chat_id, "Didn't catch that. Reply with the number or name, or /cancel.")
            return
        name, amt, row_idx = match
        addr = f"C{block['start_row'] + row_idx}"
        session["step"] = "new_value"
        session["target"] = {"label": name, "addr": addr, "is_credit": False}
        send_message(chat_id, f"New amount for {name}? (currently ${amt:,.2f}, cell {addr})")
        return

    if step == "new_value":
        amount = parse_finance_amount(stripped)
        if amount is None:
            send_message(chat_id, "Didn't catch a number. Try again, or /cancel.")
            return
        target = session["target"]
        value_to_write = -abs(amount) if (target["is_credit"] and amount) else amount
        _finish_edit(chat_id, session, target["label"], target["addr"], value_to_write)
        return


def today_label():
    now = datetime.now()
    return f"{MONTH_NAMES[now.month - 1]} {now.day}"


def start_finance_session(chat_id):
    finance_sessions[chat_id] = {
        "step": "safe",
        "values": {},
        "date_display": today_label(),
    }
    send_message(chat_id, f"New finance entry for {today_label()}. Safe? (amount, or 0 to skip. /cancel to stop)")


def handle_finance_session(chat_id, text):
    session = finance_sessions[chat_id]
    stripped = text.strip()

    if stripped.lower() == "/cancel":
        del finance_sessions[chat_id]
        send_message(chat_id, "Cancelled, nothing saved.")
        return

    step = session["step"]

    if step == "confirm":
        if stripped.lower() in ("yes", "y"):
            try:
                row, sheet_id, n_rows = write_finance_block(session)
            except Exception as e:
                print("Finance sheet write failed:", e)
                send_message(chat_id, f"Couldn't save to the Sheet ({e}). Try /log again in a bit.")
                return
            del finance_sessions[chat_id]
            url = sheet_block_url(sheet_id, row, n_rows)
            send_message(
                chat_id,
                f"Saved to the sheet starting at row {row}.",
                reply_markup=sheet_link_keyboard(url),
            )
        elif stripped.lower() in ("no", "n", "cancel"):
            del finance_sessions[chat_id]
            send_message(chat_id, "Discarded, nothing saved.")
        else:
            send_message(chat_id, "Reply yes or no.")
        return

    if step == "safe":
        amount = parse_finance_amount(stripped)
        if amount is None:
            send_message(chat_id, "Didn't catch a number for Safe. Enter an amount, or 0 to skip.")
            return
        session["values"]["Safe"] = amount
        session["step"] = "drawer"
        send_message(chat_id, "Drawer? (amount, or 0 to skip)")
        return

    if step == "drawer":
        amount = parse_finance_amount(stripped)
        if amount is None:
            send_message(chat_id, "Didn't catch a number for Drawer. Enter an amount, or 0 to skip.")
            return
        session["values"]["Drawer"] = amount
        session["step"] = "others"
        send_message(
            chat_id,
            "Anyone else? List amount + name pairs in one message, e.g.\n"
            "299 george -100 bassel -144 jose 350 jaber\n"
            "(positive = they owe you, negative = you owe them). Or reply no/done.",
        )
        return

    if step == "others":
        if stripped.lower() in ("no", "none", "done", "skip", "-"):
            items = []
        else:
            items = parse_others(stripped)
        session["others_items"] = items
        session["values"]["Others"] = sum(amt for amt, _ in items)
        session["others_detail"] = format_others_detail(items)
        session["step"] = FINANCE_BANK_ORDER[0]
        first_bank = FINANCE_BANK_ORDER[0]
        send_message(
            chat_id, f"{first_bank}? (amount, or 0 to skip)", reply_markup=bank_keyboard(first_bank)
        )
        return

    if step in FINANCE_BANK_ORDER:
        amount = parse_finance_amount(stripped)
        if amount is None:
            send_message(chat_id, f"Didn't catch a number for {step}. Enter an amount, or 0 to skip.")
            return
        session["values"][step] = amount
        idx = FINANCE_BANK_ORDER.index(step)
        if idx + 1 < len(FINANCE_BANK_ORDER):
            next_bank = FINANCE_BANK_ORDER[idx + 1]
            session["step"] = next_bank
            send_message(
                chat_id, f"{next_bank}? (amount, or 0 to skip)", reply_markup=bank_keyboard(next_bank)
            )
            return

        first_key, first_label = FINANCE_CREDIT_ORDER[0]
        session["step"] = first_key
        send_message(chat_id, f"{first_label}? (amount owed, or 0 to skip)")
        return

    credit_keys = [k for k, _ in FINANCE_CREDIT_ORDER]
    credit_labels = dict(FINANCE_CREDIT_ORDER)

    if step in credit_keys:
        amount = parse_finance_amount(stripped)
        if amount is None:
            send_message(
                chat_id,
                f"Didn't catch a number for {credit_labels[step]}. Enter the amount owed, or 0 to skip.",
            )
            return
        # Stored as negative: these balances reduce the total.
        session["values"][step] = -abs(amount) if amount else 0.0
        idx = credit_keys.index(step)
        if idx + 1 < len(credit_keys):
            next_key = credit_keys[idx + 1]
            session["step"] = next_key
            send_message(chat_id, f"{credit_labels[next_key]}? (amount owed, or 0 to skip)")
            return

        pap, _ = get_total()
        session["pap"] = pap
        others_sum = session["values"]["Others"]
        subtotal = (
            session["values"]["Safe"]
            + session["values"]["Drawer"]
            + others_sum
            + sum(session["values"][b] for b in FINANCE_BANK_ORDER)
            + sum(session["values"][k] for k in credit_keys)
        )
        net = subtotal - pap
        session["step"] = "confirm"
        summary = (
            f"Date: {session['date_display']}\n"
            f"Safe: ${session['values']['Safe']:,.2f}\n"
            f"Drawer: ${session['values']['Drawer']:,.2f}\n"
            f"Others: {session['others_detail']} (net ${others_sum:,.2f})\n"
            + "\n".join(f"{b}: ${session['values'][b]:,.2f}" for b in FINANCE_BANK_ORDER)
            + "\n"
            + "\n".join(f"{credit_labels[k]}: ${session['values'][k]:,.2f}" for k in credit_keys)
            + f"\n\nSubtotal: ${subtotal:,.2f}"
            + f"\nPAP (dad's money): ${pap:,.2f}"
            + f"\nNet total: ${net:,.2f}"
            + "\n\nSave this to the sheet? yes/no"
        )
        send_message(chat_id, summary)
        return


# Recognized ways someone might mention a currency, mapped to its code.
# Longer phrases are checked before shorter ones so "us dollars" matches
# before a bare "us" would.
CURRENCY_ALIASES = {
    "us dollars": "USD", "us dollar": "USD", "american dollars": "USD",
    "american dollar": "USD", "usd": "USD",
    "canadian dollars": "CAD", "canadian dollar": "CAD", "cad": "CAD",
    "euros": "EUR", "euro": "EUR", "eur": "EUR",
    "pounds": "GBP", "pound sterling": "GBP", "gbp": "GBP",
    "dirhams": "AED", "dirham": "AED", "aed": "AED",
    "riyals": "SAR", "riyal": "SAR", "sar": "SAR",
    "lebanese pounds": "LBP", "lebanese pound": "LBP", "lbp": "LBP",
    "jod": "JOD", "egp": "EGP", "chf": "CHF", "qar": "QAR", "kwd": "KWD",
}
CURRENCY_PHRASES = sorted(CURRENCY_ALIASES.items(), key=lambda kv: -len(kv[0]))

# --- Excel styling ---
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_SIDE = Side(style="thin", color="B7B7B7")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

pending_confirmations = {}

HELP_TEXT = (
    "Hey! Just text me an amount whenever your dad sends you money, e.g.:\n\n"
    "  5000 from Bassam Hanna\n"
    "  five thousand dollars from Serge\n"
    "  5000 US dollars from Bassam Hanna   (auto-converts to CAD)\n\n"
    "Commands:\n"
    "/total - see your running total\n"
    "/thismonth - total for the current month\n"
    "/bysender - breakdown by who sent it\n"
    "/download - get the Excel file\n"
    "/undo - remove the last entry\n"
    "/log - log today's numbers into the Joe Finance Tracker Sheet\n"
    "/edit - fix a number on the most recent entry (Subtotal/Net recalculate automatically)\n\n"
    "Edited the Excel file yourself? Just send it back to me as a file "
    "attachment and I'll use your edited version going forward.\n"
)

AMOUNT_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")

NUMBER_WORDS = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16, "seventeen": 17,
    "eighteen": 18, "nineteen": 19, "twenty": 20, "thirty": 30, "forty": 40,
    "fifty": 50, "sixty": 60, "seventy": 70, "eighty": 80, "ninety": 90,
}
SCALE_WORDS = {"hundred": 100, "thousand": 1000, "million": 1000000, "billion": 1000000000}
FILLER_AFTER_AMOUNT = {"dollars", "dollar", "usd", "bucks"}


def words_to_number(words):
    total = 0
    current = 0
    for w in words:
        if w in NUMBER_WORDS:
            current += NUMBER_WORDS[w]
        elif w in SCALE_WORDS:
            scale = SCALE_WORDS[w]
            current = (current or 1) * scale
            if scale >= 1000:
                total += current
                current = 0
    return total + current


def clean_word(w):
    return re.sub(r"[^a-zA-Z]", "", w).lower()


def extract_amount_and_note(text):
    # 1) Prefer plain digits, e.g. "5000 from Bassam Hanna"
    match = AMOUNT_RE.search(text)
    if match:
        amount = float(match.group())
        note = (text[:match.start()] + text[match.end():]).strip(" $-")
        return amount, note

    # 2) Fall back to spelled-out numbers, e.g. "five thousand dollars from Bassam Hanna"
    words = text.split()
    n = len(words)
    best_start = best_end = None
    i = 0
    while i < n:
        if clean_word(words[i]) in NUMBER_WORDS or clean_word(words[i]) in SCALE_WORDS:
            start = i
            j = i
            while j < n and (clean_word(words[j]) in NUMBER_WORDS or clean_word(words[j]) in SCALE_WORDS):
                j += 1
            if best_start is None or (j - start) > (best_end - best_start):
                best_start, best_end = start, j
            i = j
        else:
            i += 1

    if best_start is None:
        return None, text

    number_words = [clean_word(w) for w in words[best_start:best_end]]
    amount = words_to_number(number_words)
    if amount <= 0:
        return None, text

    remaining = words[:best_start] + words[best_end:]
    if remaining and clean_word(remaining[0]) in FILLER_AFTER_AMOUNT:
        remaining = remaining[1:]
    note = " ".join(remaining).strip(" $-")
    return amount, note


def split_currency(note):
    """If note starts with a recognized currency word/phrase, split it out."""
    lower = note.lower()
    for phrase, code in CURRENCY_PHRASES:
        if lower == phrase or lower.startswith(phrase + " "):
            return code, note[len(phrase):].strip()
    return None, note


def convert_currency(amount, from_code):
    if from_code == HOME_CURRENCY:
        return amount, None
    try:
        resp = requests.get(f"https://open.er-api.com/v6/latest/{from_code}", timeout=6)
        data = resp.json()
        if data.get("result") != "success":
            return None, "conversion service unavailable"
        rate = data.get("rates", {}).get(HOME_CURRENCY)
        if rate is None:
            return None, f"no rate for {HOME_CURRENCY}"
        return amount * rate, None
    except Exception as e:
        return None, str(e)


def is_total_row(values):
    return bool(values) and values[0] is not None and str(values[0]).strip().upper() == "TOTAL"


def get_last_entry(ws):
    for r in range(ws.max_row, 1, -1):
        values = [c.value for c in ws[r]]
        if is_total_row(values):
            continue
        return values
    return None


def is_recent_duplicate(amount, minutes=10):
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    last = get_last_entry(ws)
    if not last or last[1] is None:
        return False
    try:
        last_dt = datetime.strptime(str(last[0]), "%Y-%m-%d %H:%M")
    except ValueError:
        return False
    same_amount = abs(float(last[1]) - float(amount)) < 0.01
    recent = (datetime.now() - last_dt).total_seconds() <= minutes * 60
    return same_amount and recent


def extract_sender(note):
    m = re.search(r"from\s+(.+)", note or "", re.IGNORECASE)
    if not m:
        return "Unknown"
    name = re.split(r"\s*\(", m.group(1).strip())[0].strip()
    return name.title() if name else "Unknown"


def get_sender_breakdown():
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    totals = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if is_total_row(row):
            continue
        amount, note = row[1], row[2]
        if amount is None:
            continue
        totals[extract_sender(note)] += float(amount)
    return totals


def get_month_total(year, month):
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    total, count = 0.0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str, amount = row[0], row[1]
        if not date_str or amount is None:
            continue
        try:
            dt = datetime.strptime(str(date_str), "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        if dt.year == year and dt.month == month:
            total += float(amount)
            count += 1
    return total, count


def style_header(ws, ncols=3):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def apply_borders(ws, ncols=3):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ncols):
        if all(c.value is None for c in row):
            continue
        for c in row:
            c.border = CELL_BORDER


def strip_total_row(ws):
    if ws.max_row > 1:
        values = [c.value for c in ws[ws.max_row]]
        if is_total_row(values):
            ws.delete_rows(ws.max_row)


def add_log_total_row(ws):
    total = 0.0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is not None:
            total += float(row[1])
    ws.append(["TOTAL", round(total, 2), ""])
    r = ws.max_row
    for col in range(1, 4):
        c = ws.cell(row=r, column=col)
        c.font = Font(bold=True)
        c.fill = TOTAL_FILL


def finalize_log_sheet(ws):
    """Recompute the total row and re-apply styling. Call after any edit to Log."""
    strip_total_row(ws)
    if ws.max_row > 1:
        add_log_total_row(ws)
    style_header(ws)
    apply_borders(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 45


def ensure_workbook():
    if not os.path.exists(XLSX_PATH):
        if download_from_drive():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Date", "Amount", "Note"])
        style_header(ws)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 45
        wb.save(XLSX_PATH)
        upload_to_drive()


def rebuild_summary(wb):
    log_ws = wb["Log"]
    monthly = defaultdict(lambda: defaultdict(float))

    for row in log_ws.iter_rows(min_row=2, values_only=True):
        if is_total_row(row):
            continue
        date_str, amount = row[0], row[1]
        if not date_str or amount is None:
            continue
        try:
            dt = datetime.strptime(str(date_str), "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        monthly[dt.year][dt.month] += float(amount)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    ws.append(["Year", "Month", "Total"])
    style_header(ws)

    for year in sorted(monthly.keys()):
        year_total = 0.0
        for month in sorted(monthly[year].keys()):
            total = monthly[year][month]
            year_total += total
            ws.append([year, MONTH_NAMES[month - 1], round(total, 2)])
        total_row = ws.max_row + 1
        ws.append([year, "Year Total", round(year_total, 2)])
        for col in ("A", "B", "C"):
            ws[f"{col}{total_row}"].font = Font(bold=True)
            ws[f"{col}{total_row}"].fill = TOTAL_FILL
        ws.append([])

    apply_borders(ws)
    for col, width in zip("ABC", (8, 14, 14)):
        ws.column_dimensions[col].width = width


def append_entry(amount, note):
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    strip_total_row(ws)
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), amount, note])
    finalize_log_sheet(ws)
    rebuild_summary(wb)
    wb.save(XLSX_PATH)
    upload_to_drive()


def undo_last():
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    strip_total_row(ws)
    if ws.max_row <= 1:
        return None
    last_row = [c.value for c in ws[ws.max_row]]
    ws.delete_rows(ws.max_row)
    finalize_log_sheet(ws)
    rebuild_summary(wb)
    wb.save(XLSX_PATH)
    upload_to_drive()
    return last_row


def get_total():
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb["Log"]
    total = 0.0
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if is_total_row(row):
            continue
        if row[1] is not None:
            total += float(row[1])
            count += 1
    return total, count


def send_message(chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    requests.post(f"{API_URL}/sendMessage", json=payload)


def send_document(chat_id):
    ensure_workbook()
    with open(XLSX_PATH, "rb") as f:
        requests.post(
            f"{API_URL}/sendDocument",
            data={"chat_id": chat_id},
            files={"document": ("dad_money.xlsx", f)},
        )


def handle_incoming_document(chat_id, document):
    file_name = document.get("file_name", "")
    if not file_name.lower().endswith(".xlsx"):
        send_message(chat_id, "That doesn't look like an .xlsx file, so I left your data untouched.")
        return

    file_id = document.get("file_id")
    info = requests.get(f"{API_URL}/getFile", params={"file_id": file_id}).json()
    if not info.get("ok"):
        send_message(chat_id, "Couldn't fetch that file from Telegram, try sending it again.")
        return

    file_path = info["result"]["file_path"]
    file_url = f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}"
    resp = requests.get(file_url)

    tmp_path = "incoming_tmp.xlsx"
    with open(tmp_path, "wb") as f:
        f.write(resp.content)

    try:
        wb = load_workbook(tmp_path)
        if "Log" not in wb.sheetnames:
            raise ValueError("Missing Log sheet")
    except Exception:
        os.remove(tmp_path)
        send_message(chat_id, "I couldn't read that file properly (it needs a 'Log' sheet), so I kept your old data.")
        return

    ws = wb["Log"]
    finalize_log_sheet(ws)
    rebuild_summary(wb)
    wb.save(tmp_path)
    os.replace(tmp_path, XLSX_PATH)
    upload_to_drive()
    total, count = get_total()
    send_message(chat_id, f"Got it — saved your edits. Running total is now ${total:,.2f} ({count} entries). New entries will build on top of this.")


@app.route("/webhook", methods=["POST"])
def webhook():
    update = request.get_json(force=True, silent=True) or {}
    message = update.get("message", {})
    chat_id = message.get("chat", {}).get("id")

    if chat_id is None:
        return "ok"

    if chat_id != ALLOWED_CHAT_ID:
        # Ignore anyone who isn't you.
        return "ok"

    if "document" in message:
        handle_incoming_document(chat_id, message["document"])
        return "ok"

    text = (message.get("text") or "").strip()
    if not text:
        return "ok"

    if chat_id in edit_sessions:
        handle_edit_session(chat_id, text)
        return "ok"

    if chat_id in finance_sessions:
        handle_finance_session(chat_id, text)
        return "ok"

    if text == "/log":
        start_finance_session(chat_id)
        return "ok"

    if text == "/edit":
        start_edit_session(chat_id)
        return "ok"

    # Handle a pending duplicate-confirmation from the previous message
    if chat_id in pending_confirmations:
        if text.lower() in ("yes", "y"):
            pending_amount, pending_note = pending_confirmations.pop(chat_id)
            append_entry(pending_amount, pending_note)
            total, count = get_total()
            send_message(chat_id, f"Logged ${pending_amount:,.2f}. Running total: ${total:,.2f} ({count} entries).")
            return "ok"
        elif text.lower() in ("no", "n", "cancel"):
            pending_confirmations.pop(chat_id, None)
            send_message(chat_id, "Cancelled, didn't log that one.")
            return "ok"
        else:
            pending_confirmations.pop(chat_id, None)
            # fall through, treat this as a fresh message

    if text in ("/start", "/help"):
        send_message(chat_id, HELP_TEXT)
        return "ok"

    if text == "/total":
        total, count = get_total()
        send_message(chat_id, f"Total from dad: ${total:,.2f} CAD across {count} entries.")
        return "ok"

    if text == "/thismonth":
        now = datetime.now()
        total, count = get_month_total(now.year, now.month)
        send_message(chat_id, f"{MONTH_NAMES[now.month - 1]} {now.year}: ${total:,.2f} CAD across {count} entries.")
        return "ok"

    if text == "/bysender":
        totals = get_sender_breakdown()
        if not totals:
            send_message(chat_id, "No entries yet.")
            return "ok"
        lines = [f"{sender}: ${amt:,.2f} CAD" for sender, amt in sorted(totals.items(), key=lambda x: -x[1])]
        send_message(chat_id, "Totals by sender:\n" + "\n".join(lines))
        return "ok"

    if text == "/download":
        send_document(chat_id)
        return "ok"

    if text == "/undo":
        removed = undo_last()
        if removed:
            send_message(chat_id, f"Removed: {removed[0]} — ${removed[1]} ({removed[2] or 'no note'})")
        else:
            send_message(chat_id, "Nothing to undo.")
        return "ok"

    amount, note = extract_amount_and_note(text)
    if amount is None:
        send_message(chat_id, "Didn't catch an amount there. Try something like: 5000 from Bassam Hanna or five thousand dollars from Serge")
        return "ok"

    currency_code, rest_note = split_currency(note)
    if currency_code and currency_code != HOME_CURRENCY:
        converted, err = convert_currency(amount, currency_code)
        if converted is None:
            note = f"{amount:,.2f} {currency_code} {rest_note} (conversion failed: {err})".strip()
        else:
            note = f"{amount:,.2f} {currency_code} {rest_note}".strip()
            amount = converted
    else:
        note = rest_note

    if is_recent_duplicate(amount):
        pending_confirmations[chat_id] = (amount, note)
        send_message(chat_id, f"Heads up — you just logged ${amount:,.2f} in the last few minutes too. Log this one as well? Reply yes or no.")
        return "ok"

    append_entry(amount, note)
    total, count = get_total()
    send_message(chat_id, f"Logged ${amount:,.2f} CAD. Running total: ${total:,.2f} CAD ({count} entries).")
    return "ok"


@app.route("/api/total", methods=["GET"])
def api_total():
    key = request.args.get("key", "")
    if not PAP_API_KEY or not hmac.compare_digest(key, PAP_API_KEY):
        return jsonify({"error": "unauthorized"}), 401
    total, count = get_total()
    return jsonify({"total": total, "count": count, "currency": HOME_CURRENCY})


@app.route("/", methods=["GET"])
def health():
    return "Bot is running."


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
